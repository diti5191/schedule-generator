from datetime import date
import os
from tempfile import NamedTemporaryFile

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, Dict
from openpyxl import load_workbook

app = FastAPI(title="CVA Scheduler API", docs_url="/docs", redoc_url="/redoc", openapi_url="/openapi.json")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"],
)

@app.get("/health")
def health():
    return {"ok": True}

@app.get("/")
def home():
    return {"message": "CVA Scheduler API is running. Open /docs for the API."}

# ---------------- GRID CHECK (kept for mapping sanity) ----------------

class Window(BaseModel):
    start_date: date
    end_date: date

@app.post(
    "/schedules/export/xlsx",
    responses={200: {"content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}}, "description": "XLSX grid check"}},
)
def export_grid_check(window: Window):
    template_path = os.getenv("TEMPLATE_PATH", "/workspaces/schedule-generator/backend/app/templates/2026_WORKBOOK_TEMPLATE.xlsx")
    if not os.path.exists(template_path):
        return {"error": f"Template not found at {template_path}"}

    wb = load_workbook(template_path)
    ws = wb.worksheets[0]

    rows_by_day = {"Mon": (11, 13), "Tue": (16, 18), "Wed": (21, 23), "Thu": (26, 28), "Fri": (31, 33)}

    ws["B12"] = "GRID Mon NI: HH=MD_A  CH=MD_B"
    ws["B17"] = "GRID Tue NI: HH=MD_C  CH=MD_D"
    ws["B22"] = "GRID Wed NI: HH=MD_E  CH=MD_F"
    ws["B27"] = "GRID Thu NI: HH=MD_G  CH=MD_H"
    ws["B32"] = "GRID Fri NI: HH=MD_I  CH=MD_J"
    ws["G3"]  = ws["B32"].value

    ws["B14"] = "GRID Mon INT: Primary. Backup."
    ws["B19"] = "GRID Tue INT: Primary. Backup."
    ws["B24"] = "GRID Wed INT: Primary. Backup."
    ws["B29"] = "GRID Thu INT: Primary. Backup."
    ws["B34"] = "GRID Fri INT: Primary. Backup."
    ws["G6"]  = ws["B34"].value

    ws["G4"] = "GRID Sat NI: MD_K"
    ws["G5"] = "GRID Sun NI: MD_L"

    ws["H21"] = "GRID OBL AM"
    ws["H23"] = "GRID OBL PM"

    # Cooper (K) all AM/PM rows
    for day, (r_am, r_pm) in rows_by_day.items():
        ws[f"K{r_am}"] = f"GRID Cooper {day} AM MD/APN"
        ws[f"K{r_pm}"] = f"GRID Cooper {day} PM MD/APN"

    # Virtua (M/N), Stratford (AB), Cherry Hill (AC), RMC (AD), Elmer (AL), WT (AT)
    for day, (r_am, r_pm) in rows_by_day.items():
        ws[f"M{r_am}"] = f"GRID Virtua M {day} AM"; ws[f"M{r_pm}"] = f"GRID Virtua M {day} PM"
        ws[f"N{r_am}"] = f"GRID Virtua V {day} AM"; ws[f"N{r_pm}"] = f"GRID Virtua V {day} PM"
        ws[f"AB{r_am}"] = f"GRID Stratford {day} AM"; ws[f"AB{r_pm}"] = f"GRID Stratford {day} PM"
        ws[f"AC{r_am}"] = f"GRID CherryHill {day} AM"; ws[f"AC{r_pm}"] = f"GRID CherryHill {day} PM"
        ws[f"AD{r_am}"] = f"GRID RMC {day} AM MD+APN"; ws[f"AD{r_pm}"] = f"GRID RMC {day} PM MD+APN"
        ws[f"AL{r_am}"] = f"GRID Elmer {day} AM"  # PM blank by rule
        ws[f"AT{r_am}"] = f"GRID WT {day} AM: MD + APN + APN"
        ws[f"AT{r_pm}"] = f"GRID WT {day} PM: MD + APN + APN"

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name); tmp.close()
    return FileResponse(tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"GRID_CHECK_{window.start_date}_{window.end_date}.xlsx")

# ---------------- REAL WRITE FROM JSON ----------------

class CallsNI(BaseModel):
    hh: Optional[str] = None
    ch: Optional[str] = None

class CallsINT(BaseModel):
    primary: Optional[str] = None
    backup: Optional[str] = None

class DayCalls(BaseModel):
    ni: Optional[CallsNI] = None
    interventional: Optional[CallsINT] = None

class WeekendCalls(BaseModel):
    fri: Optional[str] = None   # mirrored to G3 from weekday Fri NI; kept optional
    sat: Optional[str] = None   # G4
    sun: Optional[str] = None   # G5

class Calls(BaseModel):
    mon: Optional[DayCalls] = None
    tue: Optional[DayCalls] = None
    wed: Optional[DayCalls] = None
    thu: Optional[DayCalls] = None
    fri: Optional[DayCalls] = None
    weekend: Optional[WeekendCalls] = None

class SitesBlock(BaseModel):
    K: Optional[str] = None   # Cooper Clinical (MD/APN)
    M: Optional[str] = None   # Virtua M
    N: Optional[str] = None   # Virtua V
    AB: Optional[str] = None  # Stratford
    AC: Optional[str] = None  # Cherry Hill
    AD: Optional[str] = None  # RMC (MD+APN in same cell)
    AL: Optional[str] = None  # Elmer (AM only)
    AT: Optional[str] = None  # WT (1 MD + 2 APNs)

class DaySites(BaseModel):
    AM: Optional[SitesBlock] = None
    PM: Optional[SitesBlock] = None

class Sites(BaseModel):
    Mon: Optional[DaySites] = None
    Tue: Optional[DaySites] = None
    Wed: Optional[DaySites] = None
    Thu: Optional[DaySites] = None
    Fri: Optional[DaySites] = None

class OBL(BaseModel):
    am: Optional[str] = None  # goes to H21
    pm: Optional[str] = None  # goes to H23

class ExportRequest(BaseModel):
    start_date: date
    end_date: date
    calls: Optional[Calls] = None
    sites: Optional[Sites] = None
    obl: Optional[OBL] = None

@app.post(
    "/schedules/export/from-json",
    responses={200: {"content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}},
                     "description": "XLSX export from provided JSON schedule"}}
)
def export_from_json(req: ExportRequest):
    template_path = os.getenv("TEMPLATE_PATH", "/workspaces/schedule-generator/backend/app/templates/2026_WORKBOOK_TEMPLATE.xlsx")
    if not os.path.exists(template_path):
        return {"error": f"Template not found at {template_path}"}

    wb = load_workbook(template_path)
    ws = wb.worksheets[0]

    rows_by_day = {"mon": (12, 14), "tue": (17, 19), "wed": (22, 24), "thu": (27, 29), "fri": (32, 34)}
    # ^ these are the CALL rows. For site rows we use (11,13) etc below.

    # ---- Calls (NI + INT) ----
    def set_cell(addr: str, val: Optional[str]):
        if val is not None and str(val).strip() != "":
            ws[addr] = val

    if req.calls:
        # NI: B12/17/22/27/32
        for key, row in [("mon",12),("tue",17),("wed",22),("thu",27),("fri",32)]:
            day: Optional[DayCalls] = getattr(req.calls, key)
            if day and day.ni:
                left = f"HH: {day.ni.hh}" if day.ni.hh else "HH:"
                right = f"CH: {day.ni.ch}" if day.ni.ch else "CH:"
                set_cell(f"B{row}", f"{left}   {right}")

        # INT: B14/19/24/29/34
        for key, row in [("mon",14),("tue",19),("wed",24),("thu",29),("fri",34)]:
            day = getattr(req.calls, key, None)
            if day and day.interventional:
                p = day.interventional.primary or ""
                b = day.interventional.backup or ""
                set_cell(f"B{row}", f"{p}. {b}.")

        # Weekend mirrors and NI Sat/Sun
        # Mirror weekday Friday NI to G3
        if req.calls.fri and req.calls.fri.ni:
            left = f"HH: {req.calls.fri.ni.hh}" if req.calls.fri.ni.hh else "HH:"
            right = f"CH: {req.calls.fri.ni.ch}" if req.calls.fri.ni.ch else "CH:"
            set_cell("G3", f"{left}   {right}")
        # Mirror INT Friday to G6
        if req.calls.fri and req.calls.fri.interventional:
            p = req.calls.fri.interventional.primary or ""
            b = req.calls.fri.interventional.backup or ""
            set_cell("G6", f"{p}. {b}.")
        # NI Sat/Sun
        if req.calls.weekend:
            set_cell("G4", req.calls.weekend.sat)
            set_cell("G5", req.calls.weekend.sun)

    # ---- OBL (Wed only) ----
    if req.obl:
        set_cell("H21", req.obl.am)
        set_cell("H23", req.obl.pm)

    # ---- Sites (Mon–Fri, AM/PM) ----
    site_rows_by_day = {"Mon": (11, 13), "Tue": (16, 18), "Wed": (21, 23), "Thu": (26, 28), "Fri": (31, 33)}
    cols = ["K","M","N","AB","AC","AD","AL","AT"]

    def write_sites(day_name: str, ds: Optional[DaySites]):
        if not ds: return
        r_am, r_pm = site_rows_by_day[day_name]
        if ds.AM:
            for col in cols:
                val = getattr(ds.AM, col, None)
                set_cell(f"{col}{r_am}", val)
        if ds.PM:
            for col in cols:
                val = getattr(ds.PM, col, None)
                # Elmer (AL) PM intentionally can be left blank by rule; still writing if provided
                set_cell(f"{col}{r_pm}", val)

    if req.sites:
        write_sites("Mon", req.sites.Mon)
        write_sites("Tue", req.sites.Tue)
        write_sites("Wed", req.sites.Wed)
        write_sites("Thu", req.sites.Thu)
        write_sites("Fri", req.sites.Fri)

    # Save & return
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name); tmp.close()
    return FileResponse(tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"SCHEDULE_{req.start_date}_{req.end_date}.xlsx")
